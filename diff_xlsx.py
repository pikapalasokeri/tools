#!/usr/bin/env python3

import openpyxl as pyxl
import difflib


def try_convert_to_float(value):
    try:
        ret = float(value)
    except:
        ret = value
    return ret


if __name__ == "__main__":
    wb1 = pyxl.load_workbook("left.xlsx")
    wb2 = pyxl.load_workbook("right.xlsx")
    sequenceMatcher = difflib.SequenceMatcher()

    result = pyxl.Workbook()
    for spuriousSheet in result.sheetnames:
        sheet = result[spuriousSheet]
        result.remove(sheet)

    sheetNames1 = list(wb1.sheetnames)
    sheetNames2 = list(wb2.sheetnames)

    sheet = wb1[sheetNames1[0]]

    allSheetNames = list(
        set(sheetNames1 + sheetNames2)
    )  # This does not keep the order.

    for sheetName in allSheetNames:
        activeSheet = result.create_sheet(sheetName)
        column_widths = {}
        if sheetName not in sheetNames1:
            print(f"No match for '{sheetName}'. Copying to result.")
            sheet2 = wb2[sheetName]
            for row in sheet2.rows:
                for cell in row:
                    newCell = activeSheet.cell(
                        row=cell.row, column=cell.column, value=cell.value
                    )

        elif sheetName not in sheetNames2:
            print(f"No match for '{sheetName}'. Copying to result.")
            sheet1 = wb1[sheetName]
            for row in sheet1.rows:
                for cell in row:
                    newCell = activeSheet.cell(row=row, column=col, value=cell.value)

        else:
            sheet1 = wb1[sheetName]
            sheet2 = wb2[sheetName]
            print(f"Diffing '{sheetName}'...")
            for row1, row2 in zip(sheet1.rows, sheet2.rows):
                for cell1, cell2 in zip(row1, row2):
                    # TODO(stanil): We assume that there are no "gaps" in any row or column,
                    # i.e. if we have cell1.row/col == cell2.row/col.
                    row = cell1.row
                    col = cell1.column

                    # If any one cell being compared is None, then just take one of the values.
                    # This is pretty bad. TODO(stanil): Fix this and something more reasonable.
                    if type(cell1.value) is None:
                        activeCell = activeSheet.cell(
                            row=row, column=col, value=cell2.value
                        )
                    elif type(cell2.value) is None:
                        activeCell = activeSheet.cell(
                            row=row, column=col, value=cell1.value
                        )
                    else:
                        set_color = False
                        if (
                            type(cell1.value) == float and type(cell2.value) == float
                        ) or (
                            type(cell1.value) == int and type(cell2.value) == int
                        ):  # also try converting to float.
                            diff = abs(cell2.value) - abs(
                                cell1.value
                            )  # do a comparison "closer to 0" to account for positive/negative values.
                            tolerance = 0.01
                            color = "FF000000"
                            # size = cell1.font.size
                            if abs(diff) < tolerance:
                                value = cell1.value
                            else:
                                value = f"{cell1.value:.2f} / {cell2.value:.2f} / {diff:.2f}"
                                if diff > 0.0:
                                    # Red
                                    color = "FFFF0000"
                                else:
                                    # Green
                                    color = "FF00E000"
                                set_color = True
                        else:
                            if cell1.value == cell2.value:
                                value = cell1.value
                            else:
                                value = "diff"

                        maybe_float_value = try_convert_to_float(value)
                        if type(maybe_float_value) is float:
                            new_cell_value = f"{maybe_float_value:.2f}"
                        else:
                            new_cell_value = maybe_float_value

                        activeCell = activeSheet.cell(
                            row=row, column=col, value=new_cell_value
                        )
                        if set_color:
                            activeCell.font = pyxl.styles.Font(color=color, bold=True)  # size=size

                        column_widths[col] = max(
                            column_widths.get(col, 0), len(str(activeCell.value))
                        )

        for col, width in column_widths.items():
            activeSheet.column_dimensions[
                pyxl.utils.get_column_letter(col)
            ].width = width + 1

    result.save("diff.xlsx")
